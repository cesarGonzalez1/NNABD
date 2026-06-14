from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import ProtectedError
from openpyxl import load_workbook

from web.models import (
    Discapacidad,
    FamiliaLinguistica,
    Lengua,
    ModoAdquisicionLengua,
    NivelCompetenciaOral,
    TipoContacto,
    TipoDiscapacidad,
)


def clean(value):
    if value is None:
        return ""
    return str(value).strip().replace("\xa0", " ")


def rows_from_sheet(workbook, sheet_name):
    if sheet_name not in workbook.sheetnames:
        raise CommandError(f"No existe la hoja requerida: {sheet_name}")

    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    header_index = None
    for index, row in enumerate(rows):
        if clean(row[0]).lower() == "id":
            header_index = index
            break

    if header_index is None:
        raise CommandError(f"No se encontro encabezado con Id en: {sheet_name}")

    headers = [clean(value).lower() for value in rows[header_index]]
    for row in rows[header_index + 1:]:
        values = [clean(value) for value in row]
        if not values or not values[0]:
            continue
        yield dict(zip(headers, values))


def delete_unreferenced(queryset):
    deleted = 0
    protected = 0
    for obj in queryset:
        try:
            obj.delete()
            deleted += 1
        except ProtectedError:
            protected += 1
    return deleted, protected


def delete_by_catalog_id(model, keep):
    deleted = 0
    protected = 0
    for obj in model.objects.all():
        if obj.catalogo_id not in keep:
            try:
                obj.delete()
                deleted += 1
            except ProtectedError:
                protected += 1
    return deleted, protected


def normalize_bool_from_interpreter_column(value):
    text = clean(value).lower()
    # El Excel del profesor marca "sin necesidad de interprete".
    can_declare = "sin necesidad" in text or "si" in text or "sí" in text
    requires_interpreter = bool(text) and not can_declare
    return can_declare, requires_interpreter


class Command(BaseCommand):
    help = "Carga el catalogo NNA proporcionado por el profesor desde un archivo XLSX."

    def add_arguments(self, parser):
        parser.add_argument(
            "--archivo",
            required=True,
            help="Ruta al archivo catalogos nna.xlsx",
        )
        parser.add_argument(
            "--conservar-obsoletos",
            action="store_true",
            help="No intenta borrar catalogos anteriores que ya no vienen en el Excel.",
        )

    def handle(self, *args, **options):
        path = Path(options["archivo"])
        if not path.exists():
            raise CommandError(f"El archivo no existe: {path}")

        workbook = load_workbook(path, data_only=True)
        conservar = options["conservar_obsoletos"]

        with transaction.atomic():
            counts = {
                "modo": self.load_modos(workbook, conservar),
                "contacto": self.load_contactos(workbook, conservar),
                "familia": self.load_familias(workbook, conservar),
                "lengua": self.load_lenguas(workbook, conservar),
                "discapacidad": self.load_discapacidades(workbook, conservar),
                "nivel": self.load_niveles(workbook, conservar),
            }

        for name, count in counts.items():
            self.stdout.write(f"{name}: {count} registros cargados")
        self.stdout.write(self.style.SUCCESS("Catalogos NNA cargados correctamente."))

    def load_modos(self, workbook, conservar):
        keep = set()
        count = 0
        for row in rows_from_sheet(workbook, "Modo adquisicion lengua"):
            catalog_id = row["id"]
            keep.add(catalog_id)
            ModoAdquisicionLengua.objects.update_or_create(
                clave=catalog_id,
                defaults={
                    "categoria": row["categoría"],
                    "descripcion": row["cómo se adquiere"],
                    "contexto": row["contexto típico"],
                },
            )
            count += 1
        if not conservar:
            delete_unreferenced(ModoAdquisicionLengua.objects.exclude(clave__in=keep))
        return count

    def load_contactos(self, workbook, conservar):
        keep = set()
        count = 0
        for row in rows_from_sheet(workbook, "Contacto"):
            catalog_id = row["id"]
            keep.add(catalog_id)
            contacto = (
                TipoContacto.objects.filter(clave=catalog_id).first()
                or TipoContacto.objects.filter(nombre__iexact=row["nombre"]).first()
            )
            if contacto:
                contacto.clave = catalog_id
                contacto.nombre = row["nombre"]
                contacto.descripcion = row["descripción"]
                contacto.save()
            else:
                TipoContacto.objects.create(
                    clave=catalog_id,
                    nombre=row["nombre"],
                    descripcion=row["descripción"],
                )
            count += 1
        if not conservar:
            delete_unreferenced(TipoContacto.objects.exclude(clave__in=keep))
        return count

    def load_familias(self, workbook, conservar):
        keep = set()
        count = 0
        for row in rows_from_sheet(workbook, "Familia"):
            catalog_id = int(row["id"])
            keep.add(catalog_id)
            nombre = row["familia_linguistica"]
            familia = (
                FamiliaLinguistica.objects.filter(catalogo_id=catalog_id).first()
                or FamiliaLinguistica.objects.filter(nombre__iexact=nombre).first()
            )
            if familia:
                familia.catalogo_id = catalog_id
                familia.nombre = nombre
                familia.save()
            else:
                FamiliaLinguistica.objects.create(catalogo_id=catalog_id, nombre=nombre)
            count += 1
        if not conservar:
            delete_by_catalog_id(FamiliaLinguistica, keep)
        return count

    def load_lenguas(self, workbook, conservar):
        familias = {
            familia.nombre: familia
            for familia in FamiliaLinguistica.objects.exclude(catalogo_id__isnull=True)
        }
        keep = set()
        count = 0
        for row in rows_from_sheet(workbook, "Lenguas"):
            catalog_id = int(row["id"])
            keep.add(catalog_id)
            familia = familias.get(row["familia_linguistica"])
            clave = f"PROF-{catalog_id:03d}"
            nombre = row["agrupacion_linguistica"]
            lengua = (
                Lengua.objects.filter(catalogo_id=catalog_id).first()
                or Lengua.objects.filter(clave_inali=clave).first()
            )
            defaults = {
                "catalogo_id": catalog_id,
                "familia": familia,
                "nombre": nombre,
                "clave_inali": clave,
                "es_indigena": True,
                "agrupacion_linguistica": nombre,
                "variante_linguistica": row["variante_linguistica"],
                "autodenominacion": row["autodenominacion"],
                "estado_region": row["estado_region"],
            }
            if lengua:
                for field, value in defaults.items():
                    setattr(lengua, field, value)
                lengua.save()
            else:
                Lengua.objects.create(**defaults)
            count += 1
        if not conservar:
            delete_by_catalog_id(Lengua, keep)
        return count

    def load_discapacidades(self, workbook, conservar):
        keep = set()
        count = 0
        for row in rows_from_sheet(workbook, "Discapacidad"):
            catalog_id = row["id"]
            keep.add(catalog_id)
            tipo = (
                TipoDiscapacidad.objects.filter(clave_inegi=catalog_id).first()
                or TipoDiscapacidad.objects.filter(nombre__iexact=row["nombre"]).first()
            )
            if tipo:
                tipo.clave_inegi = catalog_id
                tipo.nombre = row["nombre"]
                tipo.descripcion = row["descripción"]
                tipo.save()
            else:
                tipo = TipoDiscapacidad.objects.create(
                    clave_inegi=catalog_id,
                    nombre=row["nombre"],
                    descripcion=row["descripción"],
                )
            Discapacidad.objects.update_or_create(
                tipo=tipo,
                nombre=row["nombre"],
                defaults={"descripcion": row["descripción"]},
            )
            count += 1
        if not conservar:
            delete_unreferenced(Discapacidad.objects.exclude(tipo__clave_inegi__in=keep))
            delete_unreferenced(TipoDiscapacidad.objects.exclude(clave_inegi__in=keep))
        return count

    def load_niveles(self, workbook, conservar):
        keep = set()
        count = 0
        for row in rows_from_sheet(workbook, "Nivel de competencia oral"):
            catalog_id = row["id"]
            keep.add(catalog_id)
            puede_declarar, requiere_interprete = normalize_bool_from_interpreter_column(
                row["sí sin necesidad de intérprete."]
            )
            nivel = (
                NivelCompetenciaOral.objects.filter(clave=catalog_id).first()
                or NivelCompetenciaOral.objects.filter(nombre__iexact=row["nivel práctico"]).first()
            )
            if nivel:
                nivel.clave = catalog_id
                nivel.nombre = row["nivel práctico"]
                nivel.significado = row["significado"]
                nivel.puede_declarar = puede_declarar
                nivel.requiere_interprete = requiere_interprete
                nivel.save()
            else:
                NivelCompetenciaOral.objects.create(
                    clave=catalog_id,
                    nombre=row["nivel práctico"],
                    significado=row["significado"],
                    puede_declarar=puede_declarar,
                    requiere_interprete=requiere_interprete,
                )
            count += 1
        if not conservar:
            delete_unreferenced(NivelCompetenciaOral.objects.exclude(clave__in=keep))
        return count
